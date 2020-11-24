import pyautogui
import time
import cv2
import pandas
import csv
import json
import openpyxl
from pathlib import Path
import os
from datetime import datetime
import folium

time.sleep(2)
link="https://www.doh.wa.gov/Emergencies/COVID19/DataDashboard"
pyautogui.click(1388, 2103)
time.sleep(1)
pyautogui.hotkey("Win", "up")
time.sleep(1)
pyautogui.click(2224, 117)
pyautogui.typewrite(link)
time.sleep(1)
pyautogui.typewrite(["enter"])
time.sleep(3)
button=pyautogui.locateOnScreen("C://Users//Lucky//Desktop//MAP//link_to_click.PNG", grayscale=True, confidence=.5)
btn=pyautogui.center(button)
btn_x,btn_y=btn
pyautogui.click(btn_x, btn_y)
time.sleep(3)
#print(pyautogui.position())
pyautogui.click(x=1142, y=296)
time.sleep(10)
pyautogui.hotkey("CTRL", "w")
print("Download of data from site is successfully completed!")
time.sleep(20)

xlsx_file=Path('Cases', 'C://Users//Lucky//Downloads//PUBLIC_CDC_Event_Date_SARS.xlsx') #DO NOT FORGET TO DELETE THIS FILE AFTER
wb_obj=openpyxl.load_workbook(xlsx_file)
sheet=wb_obj.active

col_names=[]
for column in sheet.iter_cols():
    col_names.append(column[0].value)
cnty_ind=col_names.index('County')
cases_ind=col_names.index('NewPos_All')

county_names=[]
cases_num=[]
combined_cnty_cases=[]
for row in sheet.iter_rows():
    cnt_nm='County'
    county_names.append(row[cnty_ind].value)
    cases_num.append(row[cases_ind].value)
    combined_dt=row[cnty_ind].value, row[cases_ind].value
    combined_cnty_cases.append(combined_dt)
combined_cnty_cases_try=[]
d=[]
s=[]
try:
    for i in range(len(combined_cnty_cases)):
        if combined_cnty_cases[i][0] != combined_cnty_cases[i + 1][0]:
            ind=combined_cnty_cases.index(combined_cnty_cases[i+1])
            d.append(ind)
            continue
        elif combined_cnty_cases[i][0] == combined_cnty_cases[i + 1][0]:
            ind2=combined_cnty_cases.index(combined_cnty_cases[i])
            s.append(ind2)
            continue
except IndexError:
    pass
res=[]
try:
    for i in range(len(combined_cnty_cases)):
        if combined_cnty_cases[i][0] != combined_cnty_cases[i + 1][0]:
            ind=combined_cnty_cases.index(combined_cnty_cases[i+1])
            for x in range(len(d)):
                if ind == d[x]:
                    need_ind=d[x]
                    dtfr=combined_cnty_cases[need_ind:d[x+1]]
                    lst=[]
                    sm_lst=[]
                    for z in range(len(dtfr)):
                        lst.append(dtfr[z][1])
                        sm=sum(lst)
                        sm_lst.append(sm)
                        continue
                    rslt=sm_lst[-1]
                    cnty=combined_cnty_cases[i + 1][0]
                    nw1=cnty.replace(" County", "")
                    ap_need=str(nw1),int(rslt)
                    res.append(ap_need)
except IndexError:
    pass

crdts=[]
with open("C://Users//Lucky//Desktop//MAP//Data_for_Pandas.csv", newline='') as csvfile: #need to have this file with coordinates data for each county. Make sure this document exists with 4 columns named "COUNTIES","DATA","LON","LAT"!!!
    reader=csv.DictReader(csvfile)
    for j in reader:
        cnty=(j["COUNTIES"])
        cases=(j["DATA"])
        ln=(j["LON"])
        lt=(j["LAT"])
        dtcts=cnty, ln, lt
        crdts.append(dtcts)
try:
 res1 = []
 for entry in range(len(res)):
     for ct_crds in range(len(crdts)):
         if res[entry][0] == crdts[ct_crds][0]:
             dt_full=res[entry][0], res[entry][1], crdts[ct_crds][1], crdts[ct_crds][2]
             res1.append(dt_full)
except IndexError:
    pass
#print(res1)

directory = str(datetime.now().strftime("%d-%m-%Y_%I-%M-%S_%p")) #creating directory inside C://Users//Lucky//Desktop//MAP//
parent_dir = "C://Users//Lucky//Desktop//MAP//"
path = os.path.join(parent_dir, directory)
os.mkdir(path)
new_csv_file=path+'//COVID.csv'

with open(new_csv_file, 'w', newline='') as csvfile:
    filewriter=csv.writer(csvfile, delimiter=',', quoting=csv.QUOTE_MINIMAL)
    filewriter.writerow(['COUNTIES', 'DATA', 'LON', 'LAT'])
    for item in res1:
        filewriter.writerow(item)
    print("Yeay")

mx_mn_lat=[]
mx_mn=[]
cnty_list=[]
ln_lt_list=[]
with open(new_csv_file, newline='') as csvfile: #accessing new created file in directory
    reader = csv.DictReader(csvfile)
    for j in reader:
        lon = (j["LON"])
        lat = (j["LAT"])
        cnty = (j["COUNTIES"])
        dt_cn=[cnty,lon,lat]
        mx_mn.append(lon)
        mx_mn_lat.append(lat)
        cnty_list.append(cnty)
        ln_lt_list.append(dt_cn)
    #print(cnty_list)
    mn=float(min(mx_mn))
    mx=float(max(mx_mn))
    #print("btwn: ",mn,"and",mx)
    mn_lat=float(min(mx_mn_lat))
    mx_lat=float(max(mx_mn_lat))
    #print("btwn: ", mn_lat, "and", mx_lat)

df = pandas.read_json("C://Users//Lucky//Desktop//MAP//US_counties_borders.json") #File that contains USA counties border coordinates. Make sure this file exists. This is template for county coordinates. Do not delete or remove this file. !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
file=df.features
new_val=[]
for i in range(len(file)):
    cnty_nm=file[i]['properties']['name']
    for z in range(len(file[i]['geometry']['coordinates'][0])):
        new=file[i]['geometry']['coordinates'][0][z][0]
        val_lat=file[i]['geometry']['coordinates'][0][z][1]
        if type(new) == float and mn >= new >= mx and mn_lat <= val_lat <= mx_lat:
            new_val.append(file[i])
            break
        elif type(new) == list and mn >= new[0] >= mx:
            new_val.append(file[i])
            break
        elif type(new) != float:
            break
#print(new_val)
needed_counties=[]
for x in range(len(new_val)):
    for a in range(len(ln_lt_list)):
        if new_val[x]['properties']['name'] == ln_lt_list[a][0]:
            #print(new_val[x]['properties']['name'])
            #print(ln_lt_list[a][0])
            csv_val=float(ln_lt_list[a][1])
            csv_val_n=int(csv_val)
            for y in range(len(new_val[x]['geometry']['coordinates'][0])):
                jsn_valy=new_val[x]['geometry']['coordinates'][0][y][0]
                if type(jsn_valy) == float:
                    jsn_rnd=int(jsn_valy)
                elif type(jsn_valy) == list:
                    jsn_rnd=int(jsn_valy[0])
                if csv_val_n == jsn_rnd or csv_val_n == jsn_rnd+1 or csv_val_n == jsn_rnd-1:
                    #print(new_val[x])
                    needed_counties.append(new_val[x])
                    break
new_txt_file=path+'//COVID.txt'
file=open(new_txt_file, 'w+') #CREATING NEW FILE
file.write(str(needed_counties))
file.seek(0)
val=file.read()
var=val.replace("]]]}}, {", "]]]}},\n{")
new_var='{"type":"FeatureCollection","features":'  + var + "}"
file.seek(0)
file.write(new_var)

file.seek(0)
val_1=file.read()
val2=val_1.replace("\'","\"")
file.seek(0)
file.write(val2)

file.seek(0)
val3=file.read()
a=json.loads(val3)

new_json_file=path+'//COVID.json'
with open(new_json_file, "w") as write_file: #CREATING NEW FILE
    json.dump(a, write_file) #writing data as one line only


some_list=[]
with open(new_csv_file, newline='') as csvfile:
    reader = csv.DictReader(csvfile)
    for j in reader:
        cnty=(j["COUNTIES"])
        cases = int((j["DATA"]))
        df=pandas.read_json(new_json_file)
        file=df.features
        for i in range(len(file)):
            dict=file[i]['properties']
            cnties_jsn=file[i]['properties']['name']
            if  cnty == cnties_jsn:
                file[i]['properties']['CASES']=cases
                some_list.append(file[i])
                break

file=open(new_txt_file, 'w+')
file.write(str(some_list))
file.seek(0)
val=file.read()
var_a=val.replace("]]]}}, {", "]]]}},\n{")
new_var='{"type":"FeatureCollection","features":'  + var_a + "}"
var=new_var.replace("\'","\"")

file.seek(0)
file.write(var)
file.seek(0)
a_2=file.read()
a_3=json.loads(a_2)

with open(new_json_file, "w") as write_file2: #CREATING NEW FILE
    json.dump(a_3, write_file2) #writing data as one line only

data = pandas.read_csv(new_csv_file)
nm = list(data["COUNTIES"])
lat = list(data["LAT"])
lon = list(data["LON"])
cases = list(data["DATA"])

def color_producer(cases_num):
    if cases_num < 50:
        return 'green'
    elif 50 <= cases_num < 300:
        return 'orange'
    elif 300 <= cases_num < 500:
        return 'pink'
    elif 500 <= cases_num < 5000:
        return 'red'
    elif 5000 <= cases_num:
        return 'black'
    else:
        return 'white'

map = folium.Map(location=[47.034751, -122.825991], zoom_start=8, tiles="Stamen Terrain")

fg = folium.FeatureGroup(name="My Map")

for lt, ln, n, cs in zip(lat, lon, nm, cases):
    fg.add_child(folium.Marker(location=[lt, ln], popup=n + ": Cases " + str(cs), icon=folium.Icon(color=color_producer(cs)), color = 'gray', fill_opacity=0.7))

fg.add_child(folium.GeoJson(data=open(new_json_file,'r',encoding='utf-8-sig').read(),
                            style_function=lambda x: {'fillColor':'white' if x['properties']['CASES'] == 0
                            else 'green' if 0 < x['properties']['CASES'] <= 50 else 'orange' if 50 < x['properties']['CASES'] <= 300
                            else 'blue' if 300 < x['properties']['CASES'] <= 500 else 'red' if 500 < x['properties']['CASES'] <= 10000 else 'black'}))

map.add_child(fg)
new_map_file=path+'//Map_Corona_WA.html'
map.save(new_map_file)
print("DONE!")

os.rename(r'C://Users//Lucky//Downloads//PUBLIC_CDC_Event_Date_SARS.xlsx',r'C://Users//Lucky//Downloads//PUBLIC_CDC_Event_Date_SARS'+str(datetime.now().strftime("%d-%m-%Y_%I-%M-%S_%p"))+'.xlsx')